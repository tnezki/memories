#!/usr/bin/env python3
from pathlib import Path
import subprocess
import sys

REPO = Path("/Users/troynezki/Documents/GitHub/algebra")
BUILDER = REPO / "agenda" / "build_agenda.py"
INDEX = REPO / "agenda" / "index.html"


def run(*args):
    cmd = [str(x) for x in args]
    print("$ " + " ".join(cmd))
    subprocess.check_call(cmd, cwd=str(REPO))


def replace_once(text, old, new, label):
    if old not in text:
        raise SystemExit(f"Could not find expected code for: {label}. Nothing committed.")
    return text.replace(old, new, 1)


def main():
    run("git", "fetch", "origin", "main")
    run(
        "git", "restore",
        "--source=origin/main",
        "--",
        "agenda/build_agenda.py",
        "agenda/index.html",
    )
    print("Restored known-good agenda files from GitHub.")

    text = BUILDER.read_text(encoding="utf-8")

    old = '''import requests
from openpyxl import load_workbook
'''
    new = '''import requests
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
'''
    text = replace_once(text, old, new, "openpyxl coordinate helper")

    old_start = text.index("\ndef cell_info(ws_values, ws_links, row, col):\n")
    old_end = text.index("\ndef download_workbook():\n", old_start)

    resolver_code = r'''

def direct_cell_link(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target

    raw = cell.value
    if isinstance(raw, str) and raw.startswith("="):
        match = re.search(r'HYPERLINK\(\s*"([^"]+)"', raw, re.I)
        if match:
            return match.group(1)

    return None


def clean_ref(ref):
    return ref.replace("$", "").strip()


def split_ref(ref):
    match = re.fullmatch(r"([A-Z]+)(\d+)", clean_ref(ref), re.I)
    if not match:
        return None
    return column_index_from_string(match.group(1).upper()), int(match.group(2))


class WorkbookLinkResolver:
    # Resolve Student Calendar -> Teacher Calendar -> Pacing hyperlinks.

    def __init__(self, wb_values, wb_formulas):
        self.wbv = wb_values
        self.wbf = wb_formulas

        needed = ("Student Calendar", "Teacher Calendar", "Pacing")
        for name in needed:
            if name not in wb_values.sheetnames or name not in wb_formulas.sheetnames:
                raise RuntimeError(f"Missing required sheet: {name}")

        self.student_v = wb_values["Student Calendar"]
        self.student_f = wb_formulas["Student Calendar"]
        self.teacher_v = wb_values["Teacher Calendar"]
        self.teacher_f = wb_formulas["Teacher Calendar"]
        self.pacing_v = wb_values["Pacing"]
        self.pacing_f = wb_formulas["Pacing"]

        self.teacher_keys = self._build_key_rows(
            self.teacher_v,
            column=1,
            start_row=15,
            end_row=min(391, self.teacher_v.max_row),
        )
        self.pacing_keys = self._build_key_rows(
            self.pacing_v,
            column=1,
            start_row=1,
            end_row=min(500, self.pacing_v.max_row),
        )

        pacing_link_count = 0
        for row in range(1, min(500, self.pacing_f.max_row) + 1):
            for col in range(2, min(26, self.pacing_f.max_column) + 1):
                if direct_cell_link(self.pacing_f.cell(row=row, column=col)):
                    pacing_link_count += 1

        print(f"Pacing direct hyperlinks found in XLSX: {pacing_link_count}")

        if pacing_link_count < 20:
            raise RuntimeError(
                "Google XLSX export did not preserve enough direct Pacing hyperlinks. "
                "The existing GitHub agenda was left unchanged."
            )

    @staticmethod
    def _key(value):
        if value is None:
            return None
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    def _build_key_rows(self, ws, column, start_row, end_row):
        out = {}
        for row in range(start_row, end_row + 1):
            key = self._key(ws.cell(row=row, column=column).value)
            if key is not None and key not in out:
                out[key] = row
        return out

    def _value_from_ref(self, ws, ref):
        parsed = split_ref(ref)
        if not parsed:
            return None
        col, row = parsed
        return ws.cell(row=row, column=col).value

    def _teacher_cell_from_student_formula(self, formula):
        if not isinstance(formula, str):
            return None

        direct = re.fullmatch(
            r"=\s*'Teacher Calendar'!\$?([A-Z]+)\$?(\d+)\s*",
            formula,
            re.I,
        )
        if direct:
            col = column_index_from_string(direct.group(1).upper())
            row = int(direct.group(2))
            return row, col

        range_match = re.search(
            r"INDEX\(\s*'Teacher Calendar'!"
            r"\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)",
            formula,
            re.I,
        )
        if not range_match:
            return None

        col1 = column_index_from_string(range_match.group(1).upper())
        row1 = int(range_match.group(2))
        col2 = column_index_from_string(range_match.group(3).upper())
        row2 = int(range_match.group(4))

        if col1 != col2:
            return None

        small_matches = re.findall(
            r"SMALL\(.*?,\s*(\d+)\s*\)",
            formula,
            flags=re.I,
        )
        if not small_matches:
            return None

        k = int(small_matches[-1])

        nonblank_rows = []
        for row in range(row1, row2 + 1):
            value = self.teacher_v.cell(row=row, column=col1).value
            if value not in (None, "", 0):
                nonblank_rows.append(row)

        if 1 <= k <= len(nonblank_rows):
            return nonblank_rows[k - 1], col1

        return None

    def _resolve_teacher_formula(self, row, col, depth=0):
        if depth > 8:
            return None

        formula_cell = self.teacher_f.cell(row=row, column=col)

        link = direct_cell_link(formula_cell)
        if link:
            return link

        formula = formula_cell.value
        if not isinstance(formula, str) or not formula.startswith("="):
            return None

        pacing_match = re.search(
            r"vlookup\(\s*(\$?[A-Z]+\$?\d+)\s*,\s*"
            r"Pacing!\$?A\$?\d+:\$?Z\$?\d+\s*,\s*(\d+)",
            formula,
            re.I,
        )
        if pacing_match:
            lookup_ref = pacing_match.group(1)
            return_col = int(pacing_match.group(2))

            lookup_value = self._key(
                self._value_from_ref(self.teacher_v, lookup_ref)
            )
            pacing_row = self.pacing_keys.get(lookup_value)

            if pacing_row is None:
                return None

            target = self.pacing_f.cell(
                row=pacing_row,
                column=return_col,
            )
            return direct_cell_link(target)

        teacher_match = re.search(
            r"vlookup\(\s*(\$?[A-Z]+\$?\d+)\s*,\s*"
            r"\$?A\$?(\d+):\$?H\$?(\d+)\s*,\s*(\d+)",
            formula,
            re.I,
        )
        if teacher_match:
            lookup_ref = teacher_match.group(1)
            start_row = int(teacher_match.group(2))
            end_row = int(teacher_match.group(3))
            return_col = int(teacher_match.group(4))

            lookup_value = self._key(
                self._value_from_ref(self.teacher_v, lookup_ref)
            )

            target_row = self.teacher_keys.get(lookup_value)
            if target_row is None or not (start_row <= target_row <= end_row):
                target_row = None
                for candidate in range(start_row, end_row + 1):
                    value = self._key(
                        self.teacher_v.cell(row=candidate, column=1).value
                    )
                    if value == lookup_value:
                        target_row = candidate
                        break

            if target_row is None:
                return None

            target_col = return_col
            return self._resolve_teacher_formula(
                target_row,
                target_col,
                depth=depth + 1,
            )

        return None

    def student_link(self, row, col):
        student_formula_cell = self.student_f.cell(row=row, column=col)

        link = direct_cell_link(student_formula_cell)
        if link:
            return link

        source = self._teacher_cell_from_student_formula(
            student_formula_cell.value
        )
        if not source:
            return None

        teacher_row, teacher_col = source
        return self._resolve_teacher_formula(
            teacher_row,
            teacher_col,
        )


def cell_info(ws_values, ws_links, row, col, resolver=None):
    vc = ws_values.cell(row=row, column=col)
    lc = ws_links.cell(row=row, column=col)

    value = vc.value
    link = direct_cell_link(lc) or direct_cell_link(vc)

    if not link and resolver is not None:
        link = resolver.student_link(row, col)

    return safe_text(value), link


def gather_items(ws_values, ws_links, start_row, end_row, col, resolver=None):
    items = []
    for row in range(start_row, end_row + 1):
        label, url = cell_info(
            ws_values,
            ws_links,
            row,
            col,
            resolver=resolver,
        )
        if label:
            items.append((label, url))
    return items

'''

    text = text[:old_start] + resolver_code + text[old_end:]

    old = '''        wsv = wb_values[SHEET_NAME]
        wsl = wb_links[SHEET_NAME]

        current_dates = [cell_info(wsv, wsl, 2, c)[0] for c in range(1, 6)]
        current_items = [gather_items(wsv, wsl, 3, 9, c) for c in range(1, 6)]
'''
    new = '''        wsv = wb_values[SHEET_NAME]
        wsl = wb_links[SHEET_NAME]
        resolver = WorkbookLinkResolver(wb_values, wb_links)

        current_dates = [
            cell_info(wsv, wsl, 2, c, resolver=resolver)[0]
            for c in range(1, 6)
        ]
        current_items = [
            gather_items(wsv, wsl, 3, 9, c, resolver=resolver)
            for c in range(1, 6)
        ]
'''
    text = replace_once(text, old, new, "current-week resolver")

    old = '''            dates = [cell_info(wsv, wsl, date_row, c)[0] for c in range(1, 6)]
'''
    new = '''            dates = [
                cell_info(wsv, wsl, date_row, c, resolver=resolver)[0]
                for c in range(1, 6)
            ]
'''
    text = replace_once(text, old, new, "archive dates resolver")

    old = '''            items = [gather_items(wsv, wsl, date_row + 1, end_row, c) for c in range(1, 6)]
'''
    new = '''            items = [
                gather_items(
                    wsv,
                    wsl,
                    date_row + 1,
                    end_row,
                    c,
                    resolver=resolver,
                )
                for c in range(1, 6)
            ]
'''
    text = replace_once(text, old, new, "archive items resolver")

    text = text.replace(
        "text-decoration:underline;",
        "text-decoration:none;",
    )

    BUILDER.write_text(text, encoding="utf-8")
    print("Patched builder to resolve links through the Pacing source.")

    run(sys.executable, str(BUILDER))

    page = INDEX.read_text(encoding="utf-8")
    link_count = page.count('<a class="cal-link')

    known_links = [
        "1_5_2_welcome.html",
        "unit_1_warmups",
        "u1_5_demo",
        "u1_5_notes",
        "practice_set_1_5",
    ]

    missing = [x for x in known_links if x not in page]

    if link_count < 30:
        raise SystemExit(
            f"Only {link_count} clickable calendar cells were generated. "
            "Nothing committed."
        )

    if missing:
        raise SystemExit(
            "Known current-week links are missing: "
            + ", ".join(missing)
            + ". Nothing committed."
        )

    if "text-decoration:underline" in page:
        raise SystemExit("Underline CSS still exists. Nothing committed.")

    print(f"Verified {link_count} clickable calendar cells.")
    print("Verified current lesson/resource URLs.")
    print("Verified underlines removed.")

    run("git", "add", "agenda/build_agenda.py", "agenda/index.html")

    changed = subprocess.run(
        ["git", "diff", "--cached", "--quiet"],
        cwd=str(REPO),
    ).returncode != 0

    if not changed:
        print("No changes to commit.")
        return

    run("git", "commit", "-m", "Restore agenda links from Pacing")
    run("git", "push", "origin", "main")

    print()
    print("DONE")
    print("- calendar cells are clickable again")
    print("- links are recovered from the direct Pacing hyperlinks")
    print("- underlines removed")
    print("- final styling preserved")
    print("- 5-minute automation unchanged")
    print("- no Google request occurs on student devices")


if __name__ == "__main__":
    main()
